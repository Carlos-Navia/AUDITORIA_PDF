#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Verificar_CUV.py

Audita subcarpetas dentro de una ruta principal:
- Cada subcarpeta (ej. FVEA9999) debe contener:
  - RIPS: {NOMBRE_CARPETA}.json
  - CUV : {NOMBRE_CARPETA}_CUV.json
- Ademas renombra PDFs permitidos a:
  TIPO_NIT_NUMEROFACTURA.pdf

Validaciones:
- RIPS: clave "numFactura" debe ser igual al nombre de la carpeta
- CUV : clave "NumFactura" debe ser igual al nombre de la carpeta
       clave "ResultState" debe ser True
       si el JSON es {} => reportar "CUV VACIO ({})."

Salida:
- Un Excel (.xlsx) con:
  Hoja "Auditoria" (encabezados JSON)
  Hoja "PDF" (CARPETA | ARCHIVO | NUEVO NOMBRAMIENTO | ACCION)
  carpeta | rips_archivo | cuv_archivo | rips_existe | cuv_existe | estado_general
  observaciones | rips_error_json | cuv_error_json

Uso:
  python Verificar_CUV.py --ruta "C:\\ruta\\principal"
  python Verificar_CUV.py   (y luego pega la ruta por consola)
"""

import argparse
import json
import os
import re
import getpass
from datetime import datetime, date, timedelta
from typing import Any, Dict, Optional, Tuple, List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


NIT_OBLIGADO = "901011395"
PDF_ALLOWED_CODES = [
    "FEV", "HEV", "EPI", "PDX", "DQX", "RAN", "CRC", "TAP", "TNA", "FAT",
    "FMO", "OPF", "LDP", "HAU", "HAO", "HAM", "PDE"
]

AUTH_INTERVAL_DAYS = 90


def _safe_load_json(path: str) -> Tuple[Optional[Any], Optional[str]]:
    """Carga JSON y devuelve (data, error)."""
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)

        # Si es lista, toma el primer dict si aplica
        if isinstance(data, list):
            if len(data) == 0:
                return [], None
            if isinstance(data[0], dict):
                return data[0], None
            return data, None

        return data, None
    except Exception as e:
        return None, f"{type(e).__name__}: {e}"


def _is_empty_object_json(data: Any) -> bool:
    return isinstance(data, dict) and len(data) == 0


def _get_bool(value: Any) -> Optional[bool]:
    """Normaliza a bool si es bool o 'true/false' (string)."""
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        v = value.strip().lower()
        if v == "true":
            return True
        if v == "false":
            return False
    return None


def _get_auth_file_path() -> str:
    base = os.environ.get("APPDATA") or os.path.expanduser("~")
    return os.path.join(base, "verificar_cuv_auth.json")


def _load_last_auth_date() -> Optional[date]:
    try:
        with open(_get_auth_file_path(), "r", encoding="utf-8") as f:
            data = json.load(f)
        raw = data.get("last_auth")
        if not raw:
            return None
        return datetime.strptime(raw, "%Y-%m-%d").date()
    except Exception:
        return None


def _save_last_auth_date(d: date) -> None:
    data = {"last_auth": d.strftime("%Y-%m-%d")}
    with open(_get_auth_file_path(), "w", encoding="utf-8") as f:
        json.dump(data, f)


def _build_time_password(now: datetime) -> str:
    return now.strftime("%d%m%H%M")


def require_password_every_90_days() -> None:
    today = datetime.now().date()
    last_auth = _load_last_auth_date()
    if last_auth and today < (last_auth + timedelta(days=AUTH_INTERVAL_DAYS)):
        return

    for _ in range(3):
        expected = _build_time_password(datetime.now())
        ans = getpass.getpass("Ingrese la contrasena: ").strip()
        if ans == expected:
            _save_last_auth_date(today)
            return
    raise SystemExit("ERROR: Contrasena invalida.")


def _detect_pdf_code(filename: str) -> Optional[str]:
    name = os.path.basename(filename)
    m = re.match(r"^([A-Z]{3})", name.upper())
    if not m:
        return None
    return m.group(1)


def audit_and_rename_pdfs(folder_path: str, folder_name: str) -> Tuple[List[Dict[str, Any]], bool]:
    pdf_rows: List[Dict[str, Any]] = []
    errors = False

    files = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]
    pdfs = [f for f in files if f.lower().endswith(".pdf")]
    pdfs.sort()

    seen_codes: Dict[str, int] = {}

    for filename in pdfs:
        src = os.path.join(folder_path, filename)
        code = _detect_pdf_code(filename)

        if not code or code not in PDF_ALLOWED_CODES:
            errors = True
            pdf_rows.append({
                "CARPETA": folder_name,
                "ARCHIVO": filename,
                "NUEVO NOMBRAMIENTO": "",
                "ACCION": "ERROR: TIPO_NO_PERMITIDO",
            })
            continue

        seen_codes[code] = seen_codes.get(code, 0) + 1
        if seen_codes[code] > 1:
            errors = True
            expected_name = f"{code}_{NIT_OBLIGADO}_{folder_name}.pdf"
            pdf_rows.append({
                "CARPETA": folder_name,
                "ARCHIVO": filename,
                "NUEVO NOMBRAMIENTO": expected_name,
                "ACCION": "ERROR: TIPO_DUPLICADO",
            })
            continue

        expected_name = f"{code}_{NIT_OBLIGADO}_{folder_name}.pdf"
        if filename == expected_name or filename.lower() == expected_name.lower():
            pdf_rows.append({
                "CARPETA": folder_name,
                "ARCHIVO": filename,
                "NUEVO NOMBRAMIENTO": expected_name,
                "ACCION": "OK",
            })
            continue

        dest = os.path.join(folder_path, expected_name)
        if os.path.exists(dest):
            errors = True
            pdf_rows.append({
                "CARPETA": folder_name,
                "ARCHIVO": filename,
                "NUEVO NOMBRAMIENTO": expected_name,
                "ACCION": "ERROR: DESTINO_EXISTE",
            })
            continue

        os.rename(src, dest)
        pdf_rows.append({
            "CARPETA": folder_name,
            "ARCHIVO": filename,
            "NUEVO NOMBRAMIENTO": expected_name,
            "ACCION": "RENOMBRADO",
        })

    return pdf_rows, errors


def audit_subfolder(folder_path: str) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    folder_name = os.path.basename(folder_path.rstrip(os.sep))

    expected_rips = f"{folder_name}.json"
    expected_cuv = f"{folder_name}_CUV.json"

    rips_path = os.path.join(folder_path, expected_rips)
    cuv_path = os.path.join(folder_path, expected_cuv)

    result: Dict[str, Any] = {
        "carpeta": folder_name,
        "rips_archivo": expected_rips,
        "cuv_archivo": expected_cuv,
        "rips_existe": os.path.isfile(rips_path),
        "cuv_existe": os.path.isfile(cuv_path),
        "estado_general": "PENDIENTE",
        "observaciones": "",
        "rips_error_json": "",
        "cuv_error_json": "",
    }

    obs: List[str] = []
    fail_reasons: List[str] = []

    # --- RIPS ---
    if not result["rips_existe"]:
        obs.append("No existe archivo RIPS esperado.")
        fail_reasons.append("FALTA_RIPS")
    else:
        data_rips, err = _safe_load_json(rips_path)
        if err:
            result["rips_error_json"] = err
            obs.append(f"RIPS JSON inválido ({err}).")
            fail_reasons.append("RIPS_JSON_INVALIDO")
        else:
            num_factura = data_rips.get("numFactura") if isinstance(data_rips, dict) else None
            if num_factura is None:
                obs.append('RIPS sin clave "numFactura".')
                fail_reasons.append("RIPS_SIN_NUMFACTURA")
            elif str(num_factura) != folder_name:
                obs.append(f'RIPS numFactura "{num_factura}" no coincide con carpeta "{folder_name}".')
                fail_reasons.append("RIPS_NUMFACTURA_NO_COINCIDE")

    # --- CUV ---
    if not result["cuv_existe"]:
        obs.append("No existe archivo CUV esperado.")
        fail_reasons.append("FALTA_CUV")
    else:
        data_cuv, err = _safe_load_json(cuv_path)
        if err:
            result["cuv_error_json"] = err
            obs.append(f"CUV JSON inválido ({err}).")
            fail_reasons.append("CUV_JSON_INVALIDO")
        else:
            if _is_empty_object_json(data_cuv):
                obs.append("CUV VACIO ({}).")
                # No lo marco como FAIL automáticamente; se reflejará en estado_general como WARN si no hay otros FAIL
            else:
                num_factura_cuv = data_cuv.get("NumFactura") if isinstance(data_cuv, dict) else None
                rs_raw = data_cuv.get("ResultState") if isinstance(data_cuv, dict) else None
                rs = _get_bool(rs_raw)

                if num_factura_cuv is None:
                    obs.append('CUV sin clave "NumFactura".')
                    fail_reasons.append("CUV_SIN_NUMFACTURA")
                elif str(num_factura_cuv) != folder_name:
                    obs.append(f'CUV NumFactura "{num_factura_cuv}" no coincide con carpeta "{folder_name}".')
                    fail_reasons.append("CUV_NUMFACTURA_NO_COINCIDE")

                if rs_raw is None:
                    obs.append('CUV sin clave "ResultState".')
                    fail_reasons.append("CUV_SIN_RESULTSTATE")
                elif rs is not True:
                    obs.append(f'CUV ResultState "{rs_raw}" (debe ser True).')
                    fail_reasons.append("CUV_RESULTSTATE_NO_TRUE")

    # --- PDF rename audit ---
    pdf_rows, pdf_errors = audit_and_rename_pdfs(folder_path, folder_name)
    if pdf_errors:
        obs.append("Errores en PDFs.")
        fail_reasons.append("PDF_ERROR")

    # --- Estado general ---
    if fail_reasons:
        result["estado_general"] = "FAIL"
    else:
        # Si no hay fallos, pero CUV está vacío => WARN
        if result["cuv_existe"] and not result["cuv_error_json"]:
            data_cuv_check, _ = _safe_load_json(cuv_path)
            if _is_empty_object_json(data_cuv_check):
                result["estado_general"] = "WARN"
            else:
                result["estado_general"] = "OK"
        else:
            result["estado_general"] = "OK"

    result["observaciones"] = " ".join(obs).strip()
    return result, pdf_rows


def _autofit_columns(ws, max_width: int = 60) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=col).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[letter].width = min(max_len + 2, max_width)


def export_to_excel(rows: List[Dict[str, Any]], pdf_rows: List[Dict[str, Any]], output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Auditoria"

    headers = [
        "carpeta",
        "rips_archivo",
        "cuv_archivo",
        "rips_existe",
        "cuv_existe",
        "estado_general",
        "observaciones",
        "rips_error_json",
        "cuv_error_json",
    ]

    ws.append(headers)

    # Estilo encabezado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5597")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for r in rows:
        ws.append([r.get(h) for h in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Centrar booleanos y estado
    center_cols = {"rips_existe", "cuv_existe", "estado_general"}
    for col_idx, h in enumerate(headers, start=1):
        if h in center_cols:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="center")

    # Colorear estado_general
    status_col = headers.index("estado_general") + 1
    for row_idx in range(2, ws.max_row + 1):
        v = ws.cell(row=row_idx, column=status_col).value
        if v == "OK":
            fill = PatternFill("solid", fgColor="C6EFCE")
        elif v == "WARN":
            fill = PatternFill("solid", fgColor="FFEB9C")
        elif v == "FAIL":
            fill = PatternFill("solid", fgColor="FFC7CE")
        else:
            fill = None
        if fill:
            ws.cell(row=row_idx, column=status_col).fill = fill

    _autofit_columns(ws)

    # Hoja PDF
    ws_pdf = wb.create_sheet(title="PDF")
    pdf_headers = ["CARPETA", "ARCHIVO", "NUEVO NOMBRAMIENTO", "ACCION"]
    ws_pdf.append(pdf_headers)

    for col_idx in range(1, len(pdf_headers) + 1):
        cell = ws_pdf.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for r in pdf_rows:
        ws_pdf.append([r.get(h) for h in pdf_headers])

    ws_pdf.freeze_panes = "A2"
    ws_pdf.auto_filter.ref = ws_pdf.dimensions
    _autofit_columns(ws_pdf)
    wb.save(output_path)


def list_subfolders(root: str) -> List[str]:
    subfolders = []
    with os.scandir(root) as it:
        for entry in it:
            if entry.is_dir():
                subfolders.append(entry.path)
    subfolders.sort()
    return subfolders


def main() -> None:
    require_password_every_90_days()
    parser = argparse.ArgumentParser(description="Auditoría de RIPS/CUV por subcarpeta y exportación a Excel.")
    parser.add_argument("--ruta", help="Ruta principal que contiene las subcarpetas.")
    parser.add_argument("--salida", help="Ruta de salida del Excel (.xlsx). (Opcional)")
    args = parser.parse_args()

    ruta_principal = args.ruta
    if not ruta_principal:
        ruta_principal = input("Ingrese la ruta principal: ").strip().strip('"').strip("'")

    if not os.path.isdir(ruta_principal):
        raise SystemExit(f"ERROR: La ruta no existe o no es carpeta: {ruta_principal}")

    subfolders = list_subfolders(ruta_principal)
    if not subfolders:
        raise SystemExit(f"ERROR: No se encontraron subcarpetas en: {ruta_principal}")

    resultados: List[Dict[str, Any]] = []
    pdf_rows_all: List[Dict[str, Any]] = []
    for f in subfolders:
        res, pdf_rows = audit_subfolder(f)
        resultados.append(res)
        pdf_rows_all.extend(pdf_rows)

    if args.salida:
        output_path = args.salida
        if not output_path.lower().endswith(".xlsx"):
            output_path += ".xlsx"
    else:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(ruta_principal, f"Auditoria_CUV_RIPS_{ts}.xlsx")

    export_to_excel(resultados, pdf_rows_all, output_path)
    print(f"\n✅ Auditoría finalizada. Excel generado en:\n{output_path}\n")


if __name__ == "__main__":
    main()
