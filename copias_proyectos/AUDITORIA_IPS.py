import os
import re
import json
import shutil
import getpass
from datetime import datetime, date, timedelta
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import openpyxl

# ============================================================
# PARAMETROS FIJOS (sin depender del Excel)
# ============================================================
ALLOWED_PREFIXES = ["FVEA", "FVEM", "FVEP"]
NIT_OBLIGADO = "901011395"
COD_PRESTADOR_BY_PREFIX = {
    "FVEP": "190010892311",
    "FVEA": "630010180202",
    "FVEM": "170010288401",
}
METHODS_WITH_RIPS_ADJUSTMENTS = {"NUEVA_EPS", "FAMISANAR"}
METHODS_WITHOUT_AUTH_NUMBER_VALIDATION = {"COOSALUD"}

PDF_REQUIRED_ALWAYS = ["FEV", "CRC", "PDE"]
PDF_REQUIRED_ONEOF = ["HEV", "HAO", "PDX"]

RIPS_SUFFIX = ""       # {NUMFACTURA}.json
CUV_SUFFIX = "_CUV"    # {NUMFACTURA}_CUV.json

CUTOFF_DATE = datetime(2025, 9, 1)  # < 2025-09-01 = RECUPERACION ; >= = CORRIENTE

MOVE_FOLDERS = True  # al dividir: True mueve / False copia

AUTH_INTERVAL_DAYS = 30
AUTH_FILE_NAME = "auditoria_ips_auth.json"


# ============================================================
# UTILIDADES
# ============================================================
def ensure_dir(path: str) -> None:
    os.makedirs(path, exist_ok=True)

def _get_auth_file_path() -> str:
    base = os.environ.get("APPDATA") or os.path.expanduser("~")
    return os.path.join(base, AUTH_FILE_NAME)

def _load_last_auth_date() -> Optional[date]:
    try:
        with open(_get_auth_file_path(), "r", encoding="utf-8") as f:
            data = json.load(f)
        raw = data.get("last_auth")
        if not raw:
            return None
        return datetime.strptime(raw, "%Y-%m-%d").date()
    except Exception:
        return None

def _save_last_auth_date(d: date) -> None:
    data = {"last_auth": d.strftime("%Y-%m-%d")}
    with open(_get_auth_file_path(), "w", encoding="utf-8") as f:
        json.dump(data, f)

def _build_time_password(now: datetime) -> str:
    return f"{NIT_OBLIGADO}{now.strftime('%d%m%H%M')}"

def require_password_every_30_days() -> None:
    today = datetime.now().date()
    last_auth = _load_last_auth_date()
    if last_auth and today < (last_auth + timedelta(days=AUTH_INTERVAL_DAYS)):
        return

    for _ in range(3):
        expected = _build_time_password(datetime.now())
        ans = getpass.getpass("Ingrese la contrasena: ").strip()
        if ans == expected:
            _save_last_auth_date(today)
            return
    raise SystemExit("ERROR: Contrasena invalida.")

def list_subfolders(root_dir: str) -> List[str]:
    return [
        os.path.join(root_dir, d)
        for d in os.listdir(root_dir)
        if os.path.isdir(os.path.join(root_dir, d))
    ]

def validate_folder_name(folder_name: str) -> Tuple[bool, str]:
    fn = folder_name.upper().strip()
    prefix_pattern = "(" + "|".join(map(re.escape, ALLOWED_PREFIXES)) + ")"
    pattern = rf"^{prefix_pattern}\d+$"
    ok = re.match(pattern, fn) is not None
    return ok, ("OK" if ok else f"No cumple patrÃ³n de carpeta. Esperado: {pattern}")

def detect_pdf_code_flexible(filename: str) -> Optional[str]:
    upper = filename.upper()
    for code in (PDF_REQUIRED_ALWAYS + PDF_REQUIRED_ONEOF):
        if re.search(rf"(^|[_\-\.\s]){code}([_\-\.\s]|$)", upper):
            return code
    return None

def expected_pdf_name(code: str, factura: str) -> str:
    return f"{code}_{NIT_OBLIGADO}_{factura}.pdf"

def validate_pdf_name_strict(filename: str, factura: str) -> Tuple[bool, str, Optional[str]]:
    fn = filename.strip()
    upper = fn.upper()
    if not upper.endswith(".PDF"):
        return False, "ExtensiÃ³n no es .pdf", None

    code = detect_pdf_code_flexible(fn)
    if not code:
        return False, "No se pudo detectar cÃ³digo soporte (FEV/CRC/PDE/HEV/HAO/PDX)", None

    exp = expected_pdf_name(code, factura).upper()
    ok = upper == exp
    return ok, ("OK" if ok else f"Nombre esperado: {exp}"), code

def validate_json_names(files: List[str], factura: str) -> Dict[str, Any]:
    rips_expected = f"{factura}{RIPS_SUFFIX}.json"
    cuv_expected = f"{factura}{CUV_SUFFIX}.json"

    files_lower = {f.lower(): f for f in files}

    return {
        "rips_expected": rips_expected,
        "cuv_expected": cuv_expected,
        "rips_found": files_lower.get(rips_expected.lower()),
        "cuv_found": files_lower.get(cuv_expected.lower()),
        "rips_ok": rips_expected.lower() in files_lower,
        "cuv_ok": cuv_expected.lower() in files_lower,
    }

def load_json_safe(path: str) -> Tuple[bool, Optional[Any], str]:
    try:
        with open(path, "r", encoding="utf-8") as f:
            return True, json.load(f), ""
    except Exception as e:
        return False, None, str(e)

def find_all_values(obj: Any, key: str) -> List[Any]:
    results = []
    if isinstance(obj, dict):
        for k, v in obj.items():
            if k == key:
                results.append(v)
            results.extend(find_all_values(v, key))
    elif isinstance(obj, list):
        for item in obj:
            results.extend(find_all_values(item, key))
    return results

def parse_fecha_inicio_atencion(rips_json: Any) -> Tuple[Optional[datetime], str]:
    fechas = find_all_values(rips_json, "fechaInicioAtencion")
    parsed: List[datetime] = []
    for f in fechas:
        if not isinstance(f, str):
            continue
        s = f.strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
            try:
                parsed.append(datetime.strptime(s, fmt))
                break
            except Exception:
                pass
    if not parsed:
        return None, "No se encontrÃ³ o no se pudo interpretar fechaInicioAtencion"
    return min(parsed), "OK"

def determine_regimen_from_tipo_usuario(rips_json: Any) -> Tuple[Optional[str], str]:
    tipos = find_all_values(rips_json, "tipoUsuario")
    tipo = None
    for t in tipos:
        if isinstance(t, str) and t.strip() in {"01", "02", "03", "04"}:
            tipo = t.strip()
            break

    if not tipo:
        return None, "No se encontrÃ³ tipoUsuario vÃ¡lido (01-04)"

    if tipo in {"01", "02", "03"}:
        return "CONTRIBUTIVO", "OK"
    if tipo == "04":
        return "SUBSIDIADO", "OK"
    return None, "tipoUsuario inesperado"

def validate_cuv(cuv_json: Any, factura: str) -> Tuple[bool, str]:
    if not isinstance(cuv_json, dict):
        return False, "CUV no es objeto JSON (dict)"
    rs = cuv_json.get("ResultState", None)
    nf = cuv_json.get("NumFactura", None)
    if rs is not True:
        return False, "ResultState no es true"
    if str(nf).strip() != factura:
        return False, "NumFactura no coincide con la carpeta"
    return True, "OK"

def validate_rips(rips_json: Any, factura: str) -> Tuple[bool, str]:
    if not isinstance(rips_json, dict):
        return False, "RIPS no es objeto JSON (dict)"
    obligado = rips_json.get("numDocumentoIdObligado", None)
    numfact = rips_json.get("numFactura", None)
    if str(obligado).strip() != NIT_OBLIGADO:
        return False, "numDocumentoIdObligado no coincide"
    if str(numfact).strip() != factura:
        return False, "numFactura no coincide con la carpeta"
    return True, "OK"


def expected_cod_prestador_for_factura(factura: str) -> Optional[str]:
    factura_upper = str(factura).strip().upper()
    for prefix, cod in COD_PRESTADOR_BY_PREFIX.items():
        if factura_upper.startswith(prefix):
            return cod
    return None


def adjust_rips_existing_values_in_place(
    rips_json: Any, factura: str, apply_value_changes: bool
) -> Dict[str, int]:
    expected_cod = expected_cod_prestador_for_factura(factura) if apply_value_changes else None
    stats = {
        "cod_prestador_changed": 0,
        "finalidad_changed_11_to_12": 0,
        "num_autorizacion_null_count": 0,
    }

    def walk(node: Any) -> None:
        if isinstance(node, dict):
            if "codPrestador" in node and expected_cod is not None:
                if str(node.get("codPrestador")).strip() != expected_cod:
                    node["codPrestador"] = expected_cod
                    stats["cod_prestador_changed"] += 1

            if apply_value_changes and "finalidadTecnologiaSalud" in node:
                finalidad = node.get("finalidadTecnologiaSalud")
                if (isinstance(finalidad, str) and finalidad.strip() == "11") or finalidad == 11:
                    node["finalidadTecnologiaSalud"] = "12"
                    stats["finalidad_changed_11_to_12"] += 1

            if "numAutorizacion" in node and node.get("numAutorizacion") is None:
                stats["num_autorizacion_null_count"] += 1

            for value in node.values():
                walk(value)
            return

        if isinstance(node, list):
            for item in node:
                walk(item)

    walk(rips_json)
    return stats

def classify_period(fecha_dt: Optional[datetime]) -> str:
    if not fecha_dt:
        return "SIN_FECHA"
    return "RECUPERACION" if fecha_dt < CUTOFF_DATE else "CORRIENTE"

def move_or_copy_folder(src: str, dst: str) -> None:
    if MOVE_FOLDERS:
        shutil.move(src, dst)
    else:
        shutil.copytree(src, dst)


# ============================================================
# BACKUP + RENOMBRAMIENTO (backup SOLO si renombra)
# ============================================================
def backup_and_rename_pdfs(folder_path: str, factura: str, backup_root: str) -> Tuple[List[Dict[str, Any]], bool, int, int, int]:
    folder_name = os.path.basename(folder_path)
    files = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]
    pdfs = [f for f in files if f.lower().endswith(".pdf")]

    acciones: List[Dict[str, Any]] = []
    rename_conflicts = False
    pdf_corrected = 0
    pdf_not_correctible = 0
    pdf_out_of_folder = 0

    token_factura_pattern = re.compile(r"\b[A-Z]{4}\d+\b")
    codes_seen: Dict[str, int] = {}
    backup_dir: Optional[str] = None

    for old_name in pdfs:
        upper = old_name.upper()
        tokens = token_factura_pattern.findall(upper)
        if tokens and any(t != folder_name.upper() for t in tokens):
            pdf_out_of_folder += 1

        code = detect_pdf_code_flexible(old_name)
        if not code:
            rename_conflicts = True
            pdf_not_correctible += 1
            acciones.append({
                "Carpeta": folder_name,
                "Archivo Original": old_name,
                "Archivo Final": "",
                "Estado PDF": "ERROR",
                "AcciÃƒÂ³n": "NO CORREGIBLE: sin cÃƒÂ³digo"
            })
            continue

        codes_seen[code] = codes_seen.get(code, 0) + 1
        if codes_seen[code] > 1:
            rename_conflicts = True
            pdf_not_correctible += 1
            acciones.append({
                "Carpeta": folder_name,
                "Archivo Original": old_name,
                "Archivo Final": "",
                "Estado PDF": "ERROR",
                "AcciÃƒÂ³n": f"NO CORREGIBLE: duplicado {code}"
            })
            continue

        new_name = expected_pdf_name(code, factura)
        old_path = os.path.join(folder_path, old_name)
        new_path = os.path.join(folder_path, new_name)

        if old_name.upper() == new_name.upper():
            acciones.append({
                "Carpeta": folder_name,
                "Archivo Original": old_name,
                "Archivo Final": new_name,
                "Estado PDF": "OK",
                "AcciÃƒÂ³n": "OK"
            })
            continue

        if os.path.exists(new_path):
            rename_conflicts = True
            pdf_not_correctible += 1
            acciones.append({
                "Carpeta": folder_name,
                "Archivo Original": old_name,
                "Archivo Final": new_name,
                "Estado PDF": "ERROR",
                "AcciÃƒÂ³n": "NO CORREGIBLE: destino existe"
            })
            continue

        if backup_dir is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_dir = os.path.join(backup_root, timestamp, folder_name)
            ensure_dir(backup_dir)

        backup_path = os.path.join(backup_dir, old_name)
        shutil.copy2(old_path, backup_path)
        os.rename(old_path, new_path)

        pdf_corrected += 1
        acciones.append({
            "Carpeta": folder_name,
            "Archivo Original": old_name,
            "Archivo Final": new_name,
            "Estado PDF": "OK",
            "AcciÃƒÂ³n": f"RENOMBRADO (backup: {backup_path})"
        })

    return acciones, rename_conflicts, pdf_corrected, pdf_not_correctible, pdf_out_of_folder


# ============================================================
# AUDITORÃƒÂA POR CARPETA (bloquea divisiÃƒÂ³n si fecha no se lee)
# ============================================================
def audit_one_folder(folder_path: str, root_dir: str, division_method: str) -> Tuple[
    Dict[str, Any],
    List[Dict[str, Any]],
    List[Dict[str, Any]],
    Dict[str, Any],
    Optional[datetime],
    Optional[str],
    bool
]:
    folder_name = os.path.basename(folder_path).strip()
    factura = folder_name

    folder_ok, _ = validate_folder_name(folder_name)

    backup_root = os.path.join(root_dir, "BACKUP_RENOMBRES")
    pdf_rows, rename_conflicts, pdf_corrected, pdf_not_correctible, pdf_out_of_folder = backup_and_rename_pdfs(
        folder_path, factura, backup_root
    )

    files = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]
    pdf_files = [f for f in files if f.lower().endswith(".pdf")]

    pdf_codes_found = []
    pdf_strict_errors = 0
    for f in pdf_files:
        ok, _, code = validate_pdf_name_strict(f, factura)
        if not ok:
            pdf_strict_errors += 1
        if code:
            pdf_codes_found.append(code)

    missing_always = [c for c in PDF_REQUIRED_ALWAYS if c not in pdf_codes_found]
    oneof_present = [c for c in PDF_REQUIRED_ONEOF if c in pdf_codes_found]

    min_pdf = len(PDF_REQUIRED_ALWAYS) + 1
    max_pdf = len(PDF_REQUIRED_ALWAYS) + len(PDF_REQUIRED_ONEOF)
    total_pdf_ok = (min_pdf <= len(pdf_files) <= max_pdf)
    oneof_ok = (len(oneof_present) >= 1)

    errores_pdf = 0
    errores_pdf += len(missing_always)
    errores_pdf += 1 if not total_pdf_ok else 0
    errores_pdf += 1 if not oneof_ok else 0
    errores_pdf += 1 if rename_conflicts else 0
    errores_pdf += 1 if pdf_strict_errors > 0 else 0

    json_names = validate_json_names(files, factura)

    json_rows: List[Dict[str, Any]] = []
    errores_json = 0

    fecha_dt: Optional[datetime] = None
    regimen: Optional[str] = None

    # ---- RIPS ----
    if json_names["rips_ok"]:
        rips_file = json_names["rips_found"]
        ok_load, rips_json, err = load_json_safe(os.path.join(folder_path, rips_file))
        if not ok_load:
            errores_json += 1
            json_rows.append({
                "Carpeta": folder_name,
                "Archivo JSON": rips_file,
                "Tipo JSON": "RIPS",
                "Estado JSON": "ERROR",
                "OBSERVACION": f"JSON invalido: {err}"
            })
        else:
            apply_rips_changes = division_method in METHODS_WITH_RIPS_ADJUSTMENTS
            apply_num_autorizacion_validation = (
                division_method not in METHODS_WITHOUT_AUTH_NUMBER_VALIDATION
            )
            rips_stats = adjust_rips_existing_values_in_place(
                rips_json, factura, apply_value_changes=apply_rips_changes
            )
            rips_observaciones: List[str] = []

            if apply_rips_changes:
                total_changes = (
                    rips_stats["cod_prestador_changed"]
                    + rips_stats["finalidad_changed_11_to_12"]
                )
                if total_changes > 0:
                    try:
                        with open(os.path.join(folder_path, rips_file), "w", encoding="utf-8") as wf:
                            json.dump(rips_json, wf, ensure_ascii=False, indent=2)
                            wf.write("\n")
                        rips_observaciones.append(
                            "Ajustes RIPS aplicados: "
                            f"codPrestador={rips_stats['cod_prestador_changed']}, "
                            "finalidadTecnologiaSalud 11->12="
                            f"{rips_stats['finalidad_changed_11_to_12']}"
                        )
                    except Exception as save_err:
                        errores_json += 1
                        rips_observaciones.append(
                            f"No se pudieron guardar ajustes RIPS: {save_err}"
                        )

            # Validacion RIPS
            vr_ok, vr_reason = validate_rips(rips_json, factura)
            if not vr_ok:
                errores_json += 1
                rips_observaciones.append(vr_reason)

            # Fecha (si falla => ERROR y BLOQUEA division)
            fecha_dt, fecha_reason = parse_fecha_inicio_atencion(rips_json)
            if not fecha_dt:
                errores_json += 1
                rips_observaciones.append(fecha_reason)

            # Regimen
            regimen, reg_reason = determine_regimen_from_tipo_usuario(rips_json)
            if regimen not in {"CONTRIBUTIVO", "SUBSIDIADO"}:
                errores_json += 1
                rips_observaciones.append(reg_reason)

            if (
                apply_num_autorizacion_validation
                and rips_stats["num_autorizacion_null_count"] > 0
            ):
                errores_json += 1
                rips_observaciones.append(
                    "NOVEDAD: numAutorizacion en null "
                    f"({rips_stats['num_autorizacion_null_count']} ocurrencias)"
                )

            auth_number_ok = (
                rips_stats["num_autorizacion_null_count"] == 0
                if apply_num_autorizacion_validation
                else True
            )
            rips_ok = (
                vr_ok
                and bool(fecha_dt)
                and regimen in {"CONTRIBUTIVO", "SUBSIDIADO"}
                and auth_number_ok
            )

            if rips_ok:
                obs = (
                    "ok"
                    if not rips_observaciones
                    else "ok | " + " | ".join(rips_observaciones)
                )
                json_rows.append({
                    "Carpeta": folder_name,
                    "Archivo JSON": rips_file,
                    "Tipo JSON": "RIPS",
                    "Estado JSON": "OK",
                    "OBSERVACION": obs
                })
            else:
                if not rips_observaciones:
                    rips_observaciones.append("Error no especificado en validacion RIPS")
                json_rows.append({
                    "Carpeta": folder_name,
                    "Archivo JSON": rips_file,
                    "Tipo JSON": "RIPS",
                    "Estado JSON": "ERROR",
                    "OBSERVACION": " | ".join(rips_observaciones)
                })
    else:
        errores_json += 1
        json_rows.append({
            "Carpeta": folder_name,
            "Archivo JSON": json_names["rips_expected"],
            "Tipo JSON": "RIPS",
            "Estado JSON": "ERROR",
            "OBSERVACION": "No existe"
        })
    # ---- CUV ----
    if json_names["cuv_ok"]:
        cuv_file = json_names["cuv_found"]
        ok_load, cuv_json, err = load_json_safe(os.path.join(folder_path, cuv_file))
        if not ok_load:
            errores_json += 1
            json_rows.append({
                "Carpeta": folder_name,
                "Archivo JSON": cuv_file,
                "Tipo JSON": "CUV",
                "Estado JSON": "ERROR",
                "OBSERVACION": f"JSON invÃƒÂ¡lido: {err}"
            })
        else:
            vc_ok, vc_reason = validate_cuv(cuv_json, factura)
            if vc_ok:
                json_rows.append({
                    "Carpeta": folder_name,
                    "Archivo JSON": cuv_file,
                    "Tipo JSON": "CUV",
                    "Estado JSON": "OK",
                    "OBSERVACION": "ok"
                })
            else:
                errores_json += 1
                json_rows.append({
                    "Carpeta": folder_name,
                    "Archivo JSON": cuv_file,
                    "Tipo JSON": "CUV",
                    "Estado JSON": "ERROR",
                    "OBSERVACION": vc_reason
                })
    else:
        errores_json += 1
        json_rows.append({
            "Carpeta": folder_name,
            "Archivo JSON": json_names["cuv_expected"],
            "Tipo JSON": "CUV",
            "Estado JSON": "ERROR",
            "OBSERVACION": "No existe"
        })

    periodo = classify_period(fecha_dt)

    # Aceptado SOLO si:
    # - no hay errores PDF
    # - no hay errores JSON (incluye fecha obligatoria)
    # - existen RIPS y CUV
    accepted = folder_ok and (errores_pdf == 0) and (errores_json == 0) and json_names["rips_ok"] and json_names["cuv_ok"]

    accion = "OK" if accepted else "REVISAR"
    resumen_row = {
        "Carpeta": folder_name,
        "Errores PDF": int(errores_pdf),
        "PDF Corregidos": int(pdf_corrected),
        "PDF Fuera de Carpeta": int(pdf_out_of_folder),
        "PDF No Corregibles": int(pdf_not_correctible),
        "Errores JSON": int(errores_json),
        "AcciÃƒÂ³n": accion,
        "FECHA": fecha_dt.strftime("%Y-%m-%d") if fecha_dt else "",
        "Periodo": periodo,
        "REGIMEN": regimen if regimen else ""
    }

    def ok_or_noexist(exists: bool) -> str:
        return "OK" if exists else "NO EXISTE"

    oblig_row = {
        "Carpeta": folder_name,
        "CRC PDF": ok_or_noexist("CRC" in pdf_codes_found),
        "PDE PDF": ok_or_noexist("PDE" in pdf_codes_found),
        "FEV PDF": ok_or_noexist("FEV" in pdf_codes_found),
        "AL MENOS UNO (HAO/HEV/PDX)": ok_or_noexist(len([c for c in ["HAO", "HEV", "PDX"] if c in pdf_codes_found]) >= 1),
        "JSON RIPS": ok_or_noexist(json_names["rips_ok"]),
        "JSON CUV": ok_or_noexist(json_names["cuv_ok"]),
        "Cumple": "SI" if accepted else "NO",
        "Periodo": periodo
    }

    return resumen_row, pdf_rows, json_rows, oblig_row, fecha_dt, regimen, accepted


# ============================================================
# EXCEL: autofilter
# ============================================================
def apply_autofilter(path_xlsx: str) -> None:
    wb = openpyxl.load_workbook(path_xlsx)
    for name in wb.sheetnames:
        ws = wb[name]
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(ws.max_column)}{ws.max_row}"
    wb.save(path_xlsx)


# ============================================================
# DIVISION: helpers
# ============================================================
def unique_dest_path(dst_path: str) -> str:
    if not os.path.exists(dst_path):
        return dst_path
    base = dst_path
    i = 1
    while True:
        candidate = f"{base}__{i}"
        if not os.path.exists(candidate):
            return candidate
        i += 1

def move_or_copy_file(src: str, dst: str) -> None:
    ensure_dir(os.path.dirname(dst))
    dst_final = unique_dest_path(dst)
    if MOVE_FOLDERS:
        shutil.move(src, dst_final)
    else:
        shutil.copy2(src, dst_final)

def transfer_invoice_assets(
    invoice_folder: str,
    dest_supports_dir: str,
    dest_rips_dir: str,
    dest_cuv_dir: Optional[str] = None,
    recuperacion_mode_put_all_json_in_rips: bool = False
) -> None:
    folder_name = os.path.basename(invoice_folder)

    files = [f for f in os.listdir(invoice_folder) if os.path.isfile(os.path.join(invoice_folder, f))]
    pdfs = [f for f in files if f.lower().endswith(".pdf")]
    xmls = [f for f in files if f.lower().endswith(".xml")]
    jsons = [f for f in files if f.lower().endswith(".json")]

    # 1) JSONs
    for jf in jsons:
        src = os.path.join(invoice_folder, jf)

        if recuperacion_mode_put_all_json_in_rips:
            dst = os.path.join(dest_rips_dir, jf)
            move_or_copy_file(src, dst)
            continue

        upper = jf.upper()
        if upper.endswith(f"{CUV_SUFFIX}.JSON"):
            dst = os.path.join(dest_cuv_dir, jf) if dest_cuv_dir else os.path.join(dest_rips_dir, jf)
            move_or_copy_file(src, dst)
        else:
            dst = os.path.join(dest_rips_dir, jf)
            move_or_copy_file(src, dst)

    # 2) SOPORTES (carpeta factura)
    ensure_dir(dest_supports_dir)
    if MOVE_FOLDERS:
        dst_folder = unique_dest_path(os.path.join(dest_supports_dir, folder_name))
        shutil.move(invoice_folder, dst_folder)
    else:
        dst_folder = unique_dest_path(os.path.join(dest_supports_dir, folder_name))
        ensure_dir(dst_folder)
        for pf in pdfs:
            shutil.copy2(os.path.join(invoice_folder, pf), os.path.join(dst_folder, pf))
        for xf in xmls:
            shutil.copy2(os.path.join(invoice_folder, xf), os.path.join(dst_folder, xf))

# ============================================================
# DIVISION: Corriente/Recuperacion (sin preguntar)
# ============================================================
def split_folders_corriente_recuperacion(root_dir: str, move_plan: List[Tuple[str, Optional[datetime], Optional[str]]]) -> None:
    # Rutas base (sin crear carpetas hasta que sean necesarias)
    corriente_base = os.path.join(root_dir, "CORRIENTE")
    corriente_soportes = os.path.join(corriente_base, "SOPORTES")
    corriente_rips = os.path.join(corriente_base, "RIPS")
    corriente_cuv = os.path.join(corriente_base, "CUV")

    recuperacion_base = os.path.join(root_dir, "RECUPERACION")
    rec_contrib_base = os.path.join(recuperacion_base, "CONTRIBUTIVO")
    rec_subsid_base = os.path.join(recuperacion_base, "SUBSIDIADO")

    rec_contrib_soportes = os.path.join(rec_contrib_base, "SOPORTES")
    rec_contrib_rips = os.path.join(rec_contrib_base, "RIPS")
    rec_subsid_soportes = os.path.join(rec_subsid_base, "SOPORTES")
    rec_subsid_rips = os.path.join(rec_subsid_base, "RIPS")

    for folder_path, fecha_dt, regimen in move_plan:
        folder_name = os.path.basename(folder_path)

        # seguridad (con auditoria OK no deberia pasar)
        if not fecha_dt or regimen not in {"CONTRIBUTIVO", "SUBSIDIADO"}:
            review = os.path.join(root_dir, "REVISION", folder_name)
            ensure_dir(os.path.dirname(review))
            if MOVE_FOLDERS:
                shutil.move(folder_path, unique_dest_path(review))
            else:
                shutil.copytree(folder_path, unique_dest_path(review))
            continue

        if fecha_dt < CUTOFF_DATE:
            # RECUPERACION: TODOS los JSON a RIPS
            if regimen == "CONTRIBUTIVO":
                transfer_invoice_assets(
                    invoice_folder=folder_path,
                    dest_supports_dir=rec_contrib_soportes,
                    dest_rips_dir=rec_contrib_rips,
                    dest_cuv_dir=None,
                    recuperacion_mode_put_all_json_in_rips=True
                )
            else:
                transfer_invoice_assets(
                    invoice_folder=folder_path,
                    dest_supports_dir=rec_subsid_soportes,
                    dest_rips_dir=rec_subsid_rips,
                    dest_cuv_dir=None,
                    recuperacion_mode_put_all_json_in_rips=True
                )
        else:
            # CORRIENTE: separar RIPS y CUV
            transfer_invoice_assets(
                invoice_folder=folder_path,
                dest_supports_dir=corriente_soportes,
                dest_rips_dir=corriente_rips,
                dest_cuv_dir=corriente_cuv,
                recuperacion_mode_put_all_json_in_rips=False
            )

    print("OK: Separacion + distribucion SOPORTES/RIPS/CUV completada.")


# ============================================================
# DIVISION (preguntar primero)
# ============================================================
def split_folders_if_user_wants(root_dir: str, move_plan: List[Tuple[str, Optional[datetime], Optional[str]]]) -> None:
    ans = input("Desea realizar la separacion CORRIENTE/RECUPERACION y distribucion SOPORTES/RIPS/CUV? (S/N): ").strip().upper()
    if ans != "S":
        print("No se realizara separacion. Proceso finalizado.")
        return

    split_folders_corriente_recuperacion(root_dir, move_plan)

# ============================================================
# FUNCIONES DE SEPARACIÃ“N SEGÃšN LOS NUEVOS REQUISITOS
# ============================================================
def split_folders_for_nueva_eps(root_dir: str, folder_path: str, fecha_dt: Optional[datetime], regimen: Optional[str]) -> None:
    folder_name = os.path.basename(folder_path)

    if not fecha_dt or regimen not in {"CONTRIBUTIVO", "SUBSIDIADO"}:
        review = os.path.join(root_dir, "REVISION", folder_name)
        ensure_dir(os.path.dirname(review))
        if MOVE_FOLDERS:
            shutil.move(folder_path, unique_dest_path(review))
        else:
            shutil.copytree(folder_path, unique_dest_path(review))
        return

    if fecha_dt >= CUTOFF_DATE:
        corriente_base = os.path.join(root_dir, "CORRIENTE")
        corriente_soportes = os.path.join(corriente_base, "SOPORTES")
        corriente_rips = os.path.join(corriente_base, "RIPS")
        corriente_cuv = os.path.join(corriente_base, "CUV")
        ensure_dir(corriente_soportes)
        ensure_dir(corriente_rips)
        ensure_dir(corriente_cuv)

        transfer_invoice_assets(
            invoice_folder=folder_path,
            dest_supports_dir=corriente_soportes,
            dest_rips_dir=corriente_rips,
            dest_cuv_dir=corriente_cuv,
            recuperacion_mode_put_all_json_in_rips=False
        )
    else:
        recuperacion_base = os.path.join(root_dir, "RECUPERACION")
        ensure_dir(recuperacion_base)

        regimen_base = os.path.join(recuperacion_base, regimen)
        rec_soportes = os.path.join(regimen_base, "SOPORTES")
        rec_rips = os.path.join(regimen_base, "RIPS")
        ensure_dir(rec_soportes)
        ensure_dir(rec_rips)

        transfer_invoice_assets(
            invoice_folder=folder_path,
            dest_supports_dir=rec_soportes,
            dest_rips_dir=rec_rips,
            dest_cuv_dir=None,
            recuperacion_mode_put_all_json_in_rips=True
        )

def split_folders_for_coosalud(root_dir: str, folder_path: str, fecha_dt: Optional[datetime], regimen: Optional[str]) -> None:
    folder_name = os.path.basename(folder_path)

    if regimen not in {"CONTRIBUTIVO", "SUBSIDIADO"}:
        review = os.path.join(root_dir, "REVISION", folder_name)
        ensure_dir(os.path.dirname(review))
        if MOVE_FOLDERS:
            shutil.move(folder_path, unique_dest_path(review))
        else:
            shutil.copytree(folder_path, unique_dest_path(review))
        return

    if regimen == "CONTRIBUTIVO":
        contrib_base = os.path.join(root_dir, "CONTRIBUTIVO")
        contrib_soportes = os.path.join(contrib_base, "SOPORTES")
        contrib_rips = os.path.join(contrib_base, "RIPS")
        ensure_dir(contrib_soportes)
        ensure_dir(contrib_rips)

        transfer_invoice_assets(
            invoice_folder=folder_path,
            dest_supports_dir=contrib_soportes,
            dest_rips_dir=contrib_rips,
            dest_cuv_dir=None,
            recuperacion_mode_put_all_json_in_rips=True
        )

    elif regimen == "SUBSIDIADO":
        subsid_base = os.path.join(root_dir, "SUBSIDIADO")
        subsid_soportes = os.path.join(subsid_base, "SOPORTES")
        subsid_rips = os.path.join(subsid_base, "RIPS")
        ensure_dir(subsid_soportes)
        ensure_dir(subsid_rips)

        transfer_invoice_assets(
            invoice_folder=folder_path,
            dest_supports_dir=subsid_soportes,
            dest_rips_dir=subsid_rips,
            dest_cuv_dir=None,
            recuperacion_mode_put_all_json_in_rips=True
        )

def split_folders_for_famisanar(root_dir: str, folder_path: str, fecha_dt: Optional[datetime], regimen: Optional[str]) -> None:
    folder_name = os.path.basename(folder_path)

    if regimen not in {"CONTRIBUTIVO", "SUBSIDIADO"}:
        review = os.path.join(root_dir, "REVISION", folder_name)
        ensure_dir(os.path.dirname(review))
        if MOVE_FOLDERS:
            shutil.move(folder_path, unique_dest_path(review))
        else:
            shutil.copytree(folder_path, unique_dest_path(review))
        return

    if regimen == "CONTRIBUTIVO":
        dest_base = os.path.join(root_dir, "CONTRIBUTIVO")
    else:
        dest_base = os.path.join(root_dir, "SUBSIDIADO")

    ensure_dir(dest_base)
    dest_path = unique_dest_path(os.path.join(dest_base, folder_name))
    if MOVE_FOLDERS:
        shutil.move(folder_path, dest_path)
    else:
        shutil.copytree(folder_path, dest_path)

# ============================================================
# FUNCION GENERAL DE SEPARACIÃ“N
# ============================================================
def split_folders_based_on_provider(root_dir: str, folder_path: str, provider: str, fecha_dt: Optional[datetime], regimen: Optional[str]) -> None:
    if provider == "NUEVA EPS":
        split_folders_for_nueva_eps(root_dir, folder_path, fecha_dt, regimen)
    elif provider == "COOSALUD":
        split_folders_for_coosalud(root_dir, folder_path, fecha_dt, regimen)
    elif provider == "FAMISANAR":
        split_folders_for_famisanar(root_dir, folder_path, fecha_dt, regimen)


# ============================================================
# UTILIDADES DE ENTRADA PARA LA DIVISION
# ============================================================
def ask_division_method() -> str:
    print("Metodo de division:")
    print("1) NUEVA EPS (CORRIENTE: SOPORTES/RIPS/CUV; RECUPERACION: regimen->SOPORTES/RIPS)")
    print("2) FAMISANAR (solo regimen, sin separar SOPORTES/RIPS/CUV)")
    print("3) COOSALUD (regimen->SOPORTES/RIPS; SOPORTES: PDF/XML, RIPS: JSON)")
    print("4) No dividir")
    while True:
        ans = input("Seleccione 1, 2, 3 o 4: ").strip()
        if ans == "1":
            return "NUEVA_EPS"
        if ans == "2":
            return "FAMISANAR"
        if ans == "3":
            return "COOSALUD"
        if ans == "4":
            return "NONE"
        print("Opcion invalida. Use 1, 2, 3 o 4.")


def ask_move_or_copy() -> None:
    global MOVE_FOLDERS
    default = "M" if MOVE_FOLDERS else "C"
    while True:
        ans = input(f"Metodo de transferencia: M=mover / C=copiar [default {default}]: ").strip().upper()
        if not ans:
            ans = default
        if ans in {"M", "C"}:
            MOVE_FOLDERS = (ans == "M")
            return
        print("Opcion invalida. Use M o C.")


# ============================================================
# INTEGRACIÃ“N DE LA SEPARACIÃ“N EN EL PROCESO PRINCIPAL
# ============================================================
def main():
    require_password_every_30_days()
    root_dir = input("Ingrese la ruta principal (carpeta raÃ­z con subcarpetas de facturas): ").strip().strip('"')
    if not os.path.isdir(root_dir):
        raise ValueError(f"La ruta no existe o no es carpeta: {root_dir}")
    division_method = ask_division_method()

    output_xlsx = os.path.join(root_dir, "AUDITORIA_SOPORTES.xlsx")
    folders = list_subfolders(root_dir)

    resumen_rows: List[Dict[str, Any]] = []
    pdf_rows_all: List[Dict[str, Any]] = []
    json_rows_all: List[Dict[str, Any]] = []
    oblig_rows: List[Dict[str, Any]] = []

    move_plan: List[Tuple[str, Optional[datetime], Optional[str]]] = []
    accepted_all = True

    for folder_path in folders:
        resumen_row, pdf_rows, json_rows, oblig_row, fecha_dt, regimen, accepted = audit_one_folder(
            folder_path, root_dir, division_method
        )

        resumen_rows.append(resumen_row)
        pdf_rows_all.extend(pdf_rows)
        json_rows_all.extend(json_rows)
        oblig_rows.append(oblig_row)

        move_plan.append((folder_path, fecha_dt, regimen))

        if not accepted:
            accepted_all = False


    df_resumen = pd.DataFrame(resumen_rows, columns=[
        "Carpeta", "Errores PDF", "PDF Corregidos", "PDF Fuera de Carpeta", "PDF No Corregibles",
        "Errores JSON", "AcciÃ³n", "FECHA", "Periodo", "REGIMEN"
    ])

    df_pdf = pd.DataFrame(pdf_rows_all, columns=[
        "Carpeta", "Archivo Original", "Archivo Final", "Estado PDF", "AcciÃ³n"
    ])

    df_json = pd.DataFrame(json_rows_all, columns=[
        "Carpeta", "Archivo JSON", "Tipo JSON", "Estado JSON", "OBSERVACION"
    ])

    df_oblig = pd.DataFrame(oblig_rows, columns=[
        "Carpeta", "CRC PDF", "PDE PDF", "FEV PDF", "AL MENOS UNO (HAO/HEV/PDX)",
        "JSON RIPS", "JSON CUV", "Cumple", "Periodo"
    ])

    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        df_resumen.to_excel(writer, sheet_name="RESUMEN", index=False)
        df_pdf.to_excel(writer, sheet_name="PDF", index=False)
        df_json.to_excel(writer, sheet_name="JSON", index=False)
        df_oblig.to_excel(writer, sheet_name="OBLIGATORIOS", index=False)

    apply_autofilter(output_xlsx)
    print(f"\nâœ… AuditorÃ­a terminada. Excel generado en:\n{output_xlsx}")

    if not accepted_all:
        print("\nðŸš« Se encontraron errores. NO se ejecutarÃ¡ la divisiÃ³n de carpetas.")
        return

    if division_method == "NONE":
        print("No se realizara separacion. Proceso finalizado.")
        return

    ask_move_or_copy()

    if division_method == "NUEVA_EPS":
        split_folders_corriente_recuperacion(root_dir, move_plan)
    elif division_method == "FAMISANAR":
        for folder_path, fecha_dt, regimen in move_plan:
            split_folders_for_famisanar(root_dir, folder_path, fecha_dt, regimen)
    elif division_method == "COOSALUD":
        for folder_path, fecha_dt, regimen in move_plan:
            split_folders_for_coosalud(root_dir, folder_path, fecha_dt, regimen)



if __name__ == "__main__":
    main()

